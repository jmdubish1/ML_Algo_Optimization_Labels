import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from data_tools import math_tools as mt
import os
import openpyxl
import ml_tools.save_handler as sh
import data_tools.data_prediction_tools as dpt


def concat_excel_files(f_path, sheet_name):
    # Get a list of all .xlsx files in the folder
    excel_files = [f for f in os.listdir(f_path) if f.endswith('.xlsx')]
    all_data = []

    # Read and concatenate data from the specified sheet
    for file in excel_files:
        file_path = os.path.join(f_path, file)
        try:
            # Read the specific sheet from the Excel file
            df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
            df = df.dropna(subset=['DateTime']).reset_index(drop=True)
            df = df.drop(columns='Unnamed: 0', errors='ignore')  # Avoid errors if the column doesn't exist

            # for col in [7] + list(range(11, len(df.columns))):
            #     df.iloc[:, col] = np.array(df.iloc[:, col], dtype=np.float32)
            all_data.append(df)
        except Exception as e:
            print(f"Error reading {sheet_name} from {file}: {e}")

    # Concatenate all DataFrames
    if all_data:
        all_data = pd.concat(all_data, ignore_index=True)
        # Convert to numeric, handling errors if any column types are incorrect
        all_data.iloc[:, 1:] = all_data.iloc[:, 1:]
        return all_data
    else:
        print(f"No valid data found for sheet {sheet_name}.")
        return None


def create_lever_ratio(array_strat, array_algo, lookback, max_lever):
    """
    For each element in array_a, divide it by the minimum of all elements
    before it in array_b.

    Parameters:
    - array_a (np.ndarray): Array to perform division.
    - array_b (np.ndarray): Array to compute minimums of previous elements.

    Returns:
    - result (np.ndarray): Resulting array after division.
    """
    result = np.ones_like(array_strat, dtype=float)

    for i in range(len(array_strat)):
        if i == 0:
            result[i] = 1
        else:
            lb = lookback if i > lookback else 0
            min_strat = np.min(array_strat[lb:i])
            min_algo = np.min(array_algo[lb:i])

            if array_algo[i] == 0 or min_strat == 0 or min_algo == 0:
                result[i] = 1
            else:
                result[i] = min(max(min_algo / min_strat, 1), max_lever)

    return result


def plot_rolling_sum(df, sheet_name, param_id, f_path):
    if df is not None and not df.empty:
        # Calculate rolling sum
        rolling_df = df.copy()
        rolling_df['DateTime'] = pd.to_datetime(rolling_df['DateTime'], errors='coerce', unit='ns')
        rolling_df = rolling_df.sort_values(by='DateTime').reset_index(drop=True)
        rolling_df['PnL_algo_tot'] = rolling_df['PnL'].cumsum()
        rolling_df['PnL_Lstm_tot'] = rolling_df['PnL_Lstm'].cumsum()

        rolling_df['Maxdraw_algo'] = mt.calculate_max_drawdown(rolling_df['PnL_algo_tot'])
        rolling_df['Maxdraw_lstm'] = mt.calculate_max_drawdown(rolling_df['PnL_Lstm_tot'])

        adj_lstm_pnl = (rolling_df['PnL_Lstm_tot'].values *
                        create_lever_ratio(rolling_df['PnL_Lstm_tot'].values,
                                           rolling_df['Maxdraw_algo'].values,
                                           lookback=25,
                                           max_lever=5))

        # Create subplots
        fig, axes = plt.subplots(3, 1, sharex=False, figsize=(12, 20))

        # PnL plot
        axes[0].plot(rolling_df['DateTime'], rolling_df["PnL_algo_tot"], label="Algo PnL Tot", color='darkblue')
        axes[0].plot(rolling_df['DateTime'], rolling_df["PnL_Lstm_tot"], label="Lstm PnL Tot", color='darkred')
        axes[0].set_title('PnL')
        axes[0].legend()
        axes[0].grid(True)

        # Max Drawdown plot
        axes[1].plot(rolling_df['DateTime'], rolling_df["Maxdraw_algo"], label="Algo", color='darkblue')
        axes[1].plot(rolling_df['DateTime'], rolling_df["Maxdraw_lstm"], label="Lstm", color='darkred')
        axes[1].set_title('Max Drawdown')
        axes[1].legend()
        axes[1].grid(True)

        # Adjusted PnL plot
        axes[2].plot(rolling_df['DateTime'], rolling_df["PnL_algo_tot"], label="Algo", color='darkblue')
        axes[2].plot(rolling_df['DateTime'], adj_lstm_pnl, label="Lstm", color='darkred')
        axes[2].set_title('PnL - Adjusted for Algo Maxdraw')
        axes[2].legend()
        axes[2].grid(True)

        # Apply x-axis formatting to the entire figure
        for ax in axes:
            ax.xaxis.set_major_formatter(plt.matplotlib.dates.DateFormatter('%Y-%m-%d'))
            ax.tick_params(axis='x', rotation=45)

        # Save the plot as an image
        img_loc = os.path.join(f_path, f"{param_id}_{sheet_name}.png")
        plt.savefig(img_loc)
        print(f"Saved plot to: {img_loc}")

        plt.close(fig)

    else:
        print(f"No data available to plot for {sheet_name}.")

    return rolling_df


def save_concated_data(f_path, sheet_name, param_id, dfs, stack_row=False):
    excel_path = os.path.join(f_path, f"{param_id}_Stats_Data.xlsx")
    if os.path.exists(excel_path):
        try:
            book = openpyxl.load_workbook(excel_path)
            if not book.sheetnames:
                book.create_sheet(sheet_name, -1)
                book.active.title = sheet_name
        except Exception as e:
            print(f"Error loading workbook {excel_path}: {e}")
            return

        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            start_positions = sh.get_excel_sheet_df_positions(dfs, stack_row)
            sh.write_metrics_to_excel(writer, dfs, sheet_name, start_positions)
    else:
        try:
            sh.create_new_excel_file(excel_path, sheet_name)
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                start_positions = sh.get_excel_sheet_df_positions(dfs, stack_row)
                sh.write_metrics_to_excel(writer, dfs, sheet_name, start_positions)
        except Exception as e:
            print(f"Error creating new workbook {excel_path}: {e}")

side = 'Bull'
folder_main = r'C:\Users\jmdub\Documents\Trading\Futures\Strategy Info\Double_Candles\ATR\NQ\5min\5min_test_20years\classification_lstm'
folder_path = os.path.join(folder_main, side, "Data")

# Filter only directories
result_files = [f for f in os.listdir(folder_path) if os.path.isdir(os.path.join(folder_path, f))]
for param in result_files:
    print(f'param: {param}')
    rf = os.path.join(folder_path, param)

    agg_pnl_data = concat_excel_files(rf, f"{side}_PnL")

    # Plot the graphs
    rolling_df = plot_rolling_sum(agg_pnl_data, f"{side}_PnL", param, folder_path)
    trade_stats_algo = dpt.lstm_trade_stats(rolling_df, pred_data=False)
    trade_stats_lstm = dpt.lstm_trade_stats(rolling_df, pred_data=True)
    save_concated_data(folder_path, f'{side}_Stats', param,
                       [trade_stats_algo, trade_stats_lstm], stack_row=True)
    save_concated_data(folder_path, f'{side}_PnL', param,
                       [rolling_df], stack_row=False)

