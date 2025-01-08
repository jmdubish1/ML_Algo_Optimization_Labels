import os
import pickle
from datetime import datetime, timedelta
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image
import keras
import ml_tools.loss_functions as lf
import ml_tools.general_lstm_tools as glt
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from ml_tools.process_handler import ProcessHandler


class SaveHandler:
    def __init__(self, process_handler: "ProcessHandler"):
        self.ph = process_handler
        self.ph.save_handler = self

        # Folder and file paths
        self.param_folder = ''
        self.data_folder = ''
        self.model_folder = ''
        self.model_save_path = ''
        self.previous_model_path = ''

        self.test_date = None
        self.check_create_model_folder()

    def _ensure_directories_exist(self, folders):
        """Ensure that all required directories exist."""
        for folder in folders:
            os.makedirs(folder, exist_ok=True)

    def check_create_model_folder(self):
        """Create required folders for saving models and parameters."""
        self.param_folder = self.ph.pathmgr.algo_side_loc
        self.data_folder = os.path.join(self.param_folder, "Data")
        self.model_folder = os.path.join(self.param_folder, "Models")

        self._ensure_directories_exist([self.param_folder, self.data_folder, self.model_folder])

    def set_model_train_paths_lstm(self):
        """Set paths for saving and loading models based on test dates."""
        self.test_date = self.ph.trade_data.curr_test_date
        previous_test_date = self.test_date - timedelta(days=self.ph.setup_params.test_period_days)
        self.model_save_path = os.path.join(
            self.model_folder, f"{self.ph.side}_{self.ph.paramset_id}",
            f"{self.test_date.strftime('%Y-%m-%d')}_model"
        )
        self.previous_model_path = os.path.join(
            self.model_folder, f"{self.ph.side}_{self.ph.paramset_id}",
            f"{previous_test_date.strftime('%Y-%m-%d')}_model"
        )

    def save_lstm_model(self, is_main_model=False):
        """Save the LSTM model to the designated path."""
        os.makedirs(self.model_save_path, exist_ok=True)
        model_path = os.path.join(self.model_save_path, "model.keras")
        self.ph.ml_model.model.save(model_path)
        print(f"Model saved: {model_path}")

        if is_main_model:
            main_train_path = os.path.join(self.model_folder, "main_model.keras")
            self.ph.ml_model.model.save(main_train_path)
            print(f"Main model saved: {main_train_path}")

    def load_lstm_model(self, model_path):
        """Load an LSTM model from the specified path."""
        _, class_weights = glt.get_class_weights(self.ph)
        custom_objects = {
            "prec_recall_loss": lf.weighted_prec_recall_loss(class_weights),
            "f1_loss": lf.weighted_f1_loss(class_weights),
        }
        self.ph.ml_model.model = keras.models.load_model(model_path, custom_objects=custom_objects)
        print(f"Model loaded: {model_path}")

    def save_scalers(self):
        """Save scalers used during training."""
        scalers = {
            "y_pnl_scaler": self.ph.ml_data.y_pnl_scaler,
            "intra_scaler": self.ph.ml_data.intra_scaler,
            "daily_scaler": self.ph.ml_data.daily_scaler,
        }
        os.makedirs(self.model_save_path, exist_ok=True)
        for name, scaler in scalers.items():
            path = os.path.join(self.model_save_path, f"{name}.pkl")
            with open(path, "wb") as f:
                pickle.dump(scaler, f)
            print(f"Scaler saved: {path}")

    def load_scalers(self):
        """Load scalers from the current or previous model paths."""
        paths = [
            (self.model_save_path, "current"),
            (self.previous_model_path, "previous"),
        ]
        for folder, label in paths:
            try:
                self.ph.ml_data.y_pnl_scaler = self._load_pickle(os.path.join(folder, "y_pnl_scaler.pkl"))
                self.ph.ml_data.intra_scaler = self._load_pickle(os.path.join(folder, "intra_scaler.pkl"))
                self.ph.ml_data.daily_scaler = self._load_pickle(os.path.join(folder, "daily_scaler.pkl"))
                print(f"{label.capitalize()} scalers loaded from: {folder}")
                break
            except FileNotFoundError:
                continue

    """XGBoost Model Functions"""
    def set_model_train_paths_xgb(self):
        """Set paths for saving and loading models based on test dates."""
        self.test_date = self.ph.trade_data.curr_test_date
        self.model_save_path = os.path.dirname(self.ph.pathmgr.model_path('', self.test_date, extension=''))

    def save_xgb_model(self, uniq_id):
        model_path = self.ph.pathmgr.model_path(uniq_id, self.test_date, extension='bin')
        self.ph.ml_model.model.save_model(model_path)

    def load_xgb_model(self, uniq_id):
        model_path = self.ph.pathmgr.model_path(uniq_id, self.test_date, extension='bin')
        self.ph.ml_model.model.load_model(model_path)

    def load_xgb_params(self):
        param_file = self.ph.pathmgr.xgb_param_data_path()
        param_df = pd.read_excel(param_file)
        return param_df

    @staticmethod
    def _load_pickle(path):
        with open(path, "rb") as f:
            return pickle.load(f)

    def save_xgboost_params(self, param_list):
        """Save XGBoost parameter results to an Excel file."""
        file_path = os.path.join(self.data_folder, f"{self.ph.paramset_id}_xgb_best_params.xlsx")
        os.makedirs(os.path.dirname(file_path), exist_ok=True)

        df = pd.DataFrame(param_list)
        if os.path.exists(file_path):
            existing_df = pd.read_excel(file_path)
            combined_df = pd.concat([existing_df, df]).drop_duplicates()
        else:
            combined_df = df

        combined_df.sort_values("AUC", ascending=False, inplace=True)
        combined_df.to_excel(file_path, index=False)
        print(f"XGBoost parameters saved: {file_path}")

    def save_plot_to_excel(self, side):
        """Save a plot of model results to an Excel file."""
        img_path = os.path.join(self.data_folder, "temp_img.png")
        if self.ph.ml_model.model_plot:
            self.ph.ml_model.model_plot.fig.savefig(img_path)

        if os.path.exists(self.save_file):
            workbook = openpyxl.load_workbook(self.save_file)
        else:
            workbook = openpyxl.Workbook()

        sheet_name = f"{side}_LR_Curve"
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)

        img = Image(img_path)
        sheet.add_image(img, "F2")
        workbook.save(self.save_file)
        print(f"Plot saved to Excel: {self.save_file}")

        os.remove(img_path)


def write_metrics_to_excel(writer, dfs, sheet_name, start_positions):
    """Write multiple DataFrames to specific positions in an Excel sheet."""
    for df, (startrow, startcol) in zip(dfs, start_positions):
        df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, startcol=startcol)


def create_new_excel_file(file_path, sheet_name):
    """Create a new Excel file with a specified sheet."""
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        pd.DataFrame().to_excel(writer, sheet_name=sheet_name, index=False)

